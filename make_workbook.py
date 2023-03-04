#!/usr/bin/python3

import os
import csv
import openpyxl

# output file
workbook_file = "ResultsWorkbook.xlsx"

# Get list of all csv files in current directory with the correct naming format
csv_files = []
for f in os.listdir('.'):
    if not (f.endswith('.csv') and f.startswith('FinalResults_plate')):
        continue
    csv_files.append(f)

# Check if output workbook already exists
if os.path.exists(workbook_file):
    # If it does, load the existing workbook
    workbook = openpyxl.load_workbook(workbook_file)
else:
    # Otherwise, create a new workbook
    workbook = openpyxl.Workbook()

    # Remove the default sheet created by openpyxl
    default_sheet = workbook["Sheet"]
    workbook.remove(default_sheet)

# Add each new sheet to the workbook if it does not already exist
for file in csv_files:
    # Extract isolate name and plate id from filename
    file_parts = file[:-4].split("_")
    isolate_name = file_parts[-1].upper()
    plate_id = file_parts[-2].replace("plate", "").upper()

    # Create a new sheet in the workbook
    sheet_name = f"{isolate_name} ({plate_id})"
    if sheet_name not in workbook.sheetnames:
        sheet = workbook.create_sheet(title=sheet_name)

        # Open the csv file and read in the data
        with open(file, "r") as csv_file:
            csv_reader = csv.reader(csv_file)

            # Loop through the rows in the csv file and add them to the sheet
            for row in csv_reader:
                sheet.append(row)

        print(f"Added sheet '{sheet_name}' to workbook")
    else:
        print(f"Skipped sheet '{sheet_name}' because it already exists in the workbook")

# Sort the sheets in the workbook alphabetically by isolate name
workbook._sheets.sort(key=lambda x: x.title)

# Save the workbook
workbook.save(workbook_file)

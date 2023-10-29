#!/usr/bin/python3
#
# This file is part of the GPM phenotyping scripts.
#
# Copyright (c) 2023 Jason Toney
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program. If not, see <https://www.gnu.org/licenses/>.

"""Combines multiple CSV results files into a single Excel workbook."""

import os
import csv
import sys
import openpyxl


def compile_workbook(workbook_file, csv_files):
    """
    Compiles the csv files into the workbook.
    """
    # Check if output workbook already exists
    if os.path.exists(workbook_file):
        # If it does, load the existing workbook
        workbook = openpyxl.load_workbook(workbook_file)
    else:
        # Otherwise, create a new workbook
        workbook = openpyxl.Workbook()

        # Remove the default sheet created by openpyxl
        default_sheet = workbook["Sheet"]
        workbook.remove(default_sheet)

    # Add each new sheet to the workbook if it does not already exist
    for file in csv_files:
        # Extract isolate name and plate id from filename
        file_parts = file[:-4].split("_")
        isolate_name = file_parts[-1].upper()
        plate_id = file_parts[-2].replace("plate", "").upper()

        # Create a new sheet in the workbook
        sheet_name = f"{isolate_name} ({plate_id})"
        if sheet_name not in workbook.sheetnames:
            sheet = workbook.create_sheet(title=sheet_name)

            # Open the csv file and read in the data
            with open(file, "r", encoding="utf-8") as csv_file:
                csv_reader = csv.reader(csv_file)

                # Loop through the rows  and add them to the sheet
                for row in csv_reader:
                    sheet.append(row)

            print(f"Added sheet {sheet_name} to workbook")
        else:
            print(f"Skipped sheet {sheet_name}: already in the workbook")

    # Sort the sheets in the workbook alphabetically by isolate name
    workbook.worksheets.sort(key=lambda sheet: sheet.title)

    # Save the workbook
    workbook.save(workbook_file)

    # Log the number of sheets in the workbook
    print(f"The workbookbook contains {len(workbook.sheetnames)} sheets")


def main():
    """Set path and get filenames"""
    # Change to the script's directory
    os.chdir(os.path.dirname(sys.argv[0]))

    # output file
    workbook_file = "GPMFungicideAssay_Workbook.xlsx"

    # Get list of all usable csv files in the results directory
    csv_files = []
    for file in sorted(
        os.listdir("results"),
        key=lambda x: x.split("_")[2],
    ):
        if not (file.endswith(".csv") and file.startswith("FinalResults_plate")):
            continue
        file_path = os.path.join("results", file)
        csv_files.append(file_path)

    if csv_files:
        compile_workbook(workbook_file, csv_files)
    else:
        print("No csv files found.")


if __name__ == "__main__":
    main()
